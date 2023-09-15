
from sys import argv as args
import os
import openpyxl as xl
from openpyxl.utils import get_column_letter

FILENAME = args[1]

EXECUTABLE_PATH = "./converter"
OUT_DIR = "./"

wb = xl.load_workbook(FILENAME)
sheet = wb.active

CUR_COL = 1
CUR_ROW = 1

def CurrentCell() -> str:
	return f"{get_column_letter(CUR_COL)}{CUR_ROW}"


# list of numbers to be returned
RES_NUM_VALUES = []


# cells in each square's side length
DATA_SQUARE_SIDE = 8

# number of squares per row
DATA_SQUARES_X = 10
# number of squares per column
DATA_SQUARES_Y = 8
# last row is different
DATA_SQUARES_LAST_ROW = 4


def ReadCurrentSquare():
	global CUR_COL, CUR_ROW
	for _ in range(DATA_SQUARE_SIDE):
		number = 0b0000_0000
		multiplier = 2**(DATA_SQUARE_SIDE - 1)
		for _ in range(DATA_SQUARE_SIDE):
			value = sheet[CurrentCell()].value != None
			value *= multiplier

			number += value

			multiplier >>= 1
			CUR_COL += 1

		RES_NUM_VALUES.append(number)
		
		CUR_COL -= DATA_SQUARE_SIDE
		CUR_ROW += 1

	CUR_ROW -= DATA_SQUARE_SIDE
#END


# normal rows
for square_rows in range(DATA_SQUARES_Y):
	for square_cols in range(DATA_SQUARES_X):
		ReadCurrentSquare()
		CUR_COL += DATA_SQUARE_SIDE
	CUR_COL = 1
	CUR_ROW += DATA_SQUARE_SIDE
# last row
for square_cols in range(DATA_SQUARES_LAST_ROW):
	ReadCurrentSquare()
	CUR_COL += DATA_SQUARE_SIDE


MON_NUMBER = RES_NUM_VALUES[-1]
RES_NUM_VALUES_STR = [str(i) for i in RES_NUM_VALUES]
# print(RES_NUM_VALUES_STR)

cmd = f"{EXECUTABLE_PATH} {OUT_DIR} {MON_NUMBER} {' '.join(RES_NUM_VALUES_STR[:-8])}"
print(cmd)
os.system(cmd)


wb.close()
